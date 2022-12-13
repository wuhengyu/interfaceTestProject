# -*- coding: utf-8 -*-
# @Time    : 2022/12/13 14:23
# @Author  : Walter
# @File    : ReaderExcel.py
# @License : (C)Copyright Walter
# @Desc    :
# 在 Python 中，可以使用第三方库来实现对 Excel 文件的常见操作，
# 例如读取和写入数据、操作单元格等。具体实现可以使用如 openpyxl、xlrd、xlwt 等库。

from openpyxl.reader.excel import load_workbook
import os


class ReaderExcel():
    def __init__(self, filename, sheetName):
        self.filename = filename
        self.wb = load_workbook(filename)
        # 通过工作的表名获取一个工作表对象
        self.sheet = self.wb[sheetName]
        # 获取工作表中的最大行号
        self.maxRowNum = self.sheet.max_row
        # 获取工作表中的最大列号
        self.max_column = self.sheet.max_column
        # 获取总的行
        self.row = self.sheet.max_row
        # 设置工作簿的默认样式
        self.wb.default_style = {"font": {"name": "Arial", "size": 12}}

    def readExcel(self):
        dataList = []
        try:
            for row in self.sheet.rows:
                tmpList = []
                for cell in row:
                    tmpList.append(cell.value)
                dataList.append(tmpList)
        except:
            print("%s加载失败" % self.filename)
        return dataList

    def saveExcel(self, row, text):
        try:
            self.sheet.cell(row, self.max_column, text)
            self.wb.save(self.filename)
        except:
            print("%s 保存失败, 检查是否正在打开" % self.filename)


if __name__ == "__main__":
    base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    dri_url = os.path.join(base_path + r"\config\demoTest.xlsx")
    excel = ReaderExcel(dri_url, "Sheet1")
    dataJson = excel.readExcel()
    print(dataJson)
    excel.saveExcel(40, 'pass')
